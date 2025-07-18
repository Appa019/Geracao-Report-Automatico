import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
import os
from typing import Dict, Any, List
from datetime import datetime

class ExcelGenerator:
    """Classe para gerar relatórios Excel formatados"""
    
    def __init__(self):
        self.colors = {
            'header': 'FF4472C4',
            'subheader': 'FFD5E4F7',
            'accent': 'FFF2F2F2',
            'warning': 'FFFFF2CC',
            'success': 'FFE2EFDA'
        }
    
    def generate_excel(self, analysis_results: Dict[str, Any], analysis_type: str, filename: str) -> str:
        """
        Gera arquivo Excel formatado baseado nos resultados da análise
        
        Args:
            analysis_results: Resultados da análise de IA
            analysis_type: Tipo de análise realizada
            filename: Nome do arquivo original
            
        Returns:
            Caminho para o arquivo Excel gerado
        """
        
        # Criar arquivo temporário
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_path = temp_file.name
        temp_file.close()
        
        # Criar workbook
        wb = openpyxl.Workbook()
        
        # Remover sheet padrão
        wb.remove(wb.active)
        
        # Criar sheets baseadas no tipo de análise
        if analysis_type == "Resumo Executivo":
            self._create_executive_summary_sheets(wb, analysis_results, filename)
        elif analysis_type == "Análise por Tópicos":
            self._create_topic_analysis_sheets(wb, analysis_results, filename)
        elif analysis_type == "Extração de Dados":
            self._create_data_extraction_sheets(wb, analysis_results, filename)
        elif analysis_type == "Análise de Cláusulas":
            self._create_clause_analysis_sheets(wb, analysis_results, filename)
        elif analysis_type == "Timeline de Eventos":
            self._create_timeline_sheets(wb, analysis_results, filename)
        else:
            self._create_general_analysis_sheets(wb, analysis_results, filename, analysis_type)
        
        # Salvar arquivo
        wb.save(temp_path)
        
        return temp_path
    
    def _create_executive_summary_sheets(self, wb: openpyxl.Workbook, results: Dict[str, Any], filename: str):
        """Cria sheets para resumo executivo"""
        
        # Sheet 1: Resumo Executivo
        ws_summary = wb.create_sheet("Resumo Executivo")
        
        # Cabeçalho
        self._add_header(ws_summary, f"Resumo Executivo - {filename}")
        
        row = 3
        
        # Resumo principal
        if 'executive_summary' in results:
            ws_summary.cell(row=row, column=1, value="RESUMO EXECUTIVO")
            self._style_subheader(ws_summary.cell(row=row, column=1))
            row += 1
            
            # Quebrar resumo em parágrafos
            summary_text = results['executive_summary']
            paragraphs = summary_text.split('\n') if '\n' in summary_text else [summary_text]
            
            for paragraph in paragraphs:
                if paragraph.strip():
                    ws_summary.cell(row=row, column=1, value=paragraph.strip())
                    ws_summary.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
                    row += 1
            
            row += 1
        
        # Principais descobertas
        if 'key_findings' in results:
            ws_summary.cell(row=row, column=1, value="PRINCIPAIS DESCOBERTAS")
            self._style_subheader(ws_summary.cell(row=row, column=1))
            row += 1
            
            for i, finding in enumerate(results['key_findings'], 1):
                ws_summary.cell(row=row, column=1, value=f"{i}. {finding}")
                row += 1
            
            row += 1
        
        # Métricas importantes
        if 'important_metrics' in results:
            ws_summary.cell(row=row, column=1, value="MÉTRICAS IMPORTANTES")
            self._style_subheader(ws_summary.cell(row=row, column=1))
            row += 1
            
            for metric in results['important_metrics']:
                ws_summary.cell(row=row, column=1, value=f"• {metric}")
                row += 1
            
            row += 1
        
        # Recomendações
        if 'recommendations' in results:
            ws_summary.cell(row=row, column=1, value="RECOMENDAÇÕES")
            self._style_subheader(ws_summary.cell(row=row, column=1))
            row += 1
            
            for i, rec in enumerate(results['recommendations'], 1):
                ws_summary.cell(row=row, column=1, value=f"{i}. {rec}")
                row += 1
        
        # Ajustar largura das colunas
        ws_summary.column_dimensions['A'].width = 80
        
        # Sheet 2: Conclusões
        if 'main_conclusions' in results:
            ws_conclusions = wb.create_sheet("Conclusões")
            self._add_header(ws_conclusions, "Principais Conclusões")
            
            row = 3
            for i, conclusion in enumerate(results['main_conclusions'], 1):
                ws_conclusions.cell(row=row, column=1, value=f"Conclusão {i}")
                self._style_subheader(ws_conclusions.cell(row=row, column=1))
                row += 1
                
                ws_conclusions.cell(row=row, column=1, value=conclusion)
                ws_conclusions.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
                row += 2
            
            ws_conclusions.column_dimensions['A'].width = 80
    
    def _create_topic_analysis_sheets(self, wb: openpyxl.Workbook, results: Dict[str, Any], filename: str):
        """Cria sheets para análise por tópicos"""
        
        # Sheet 1: Visão Geral dos Tópicos
        ws_overview = wb.create_sheet("Visão Geral")
        self._add_header(ws_overview, f"Análise por Tópicos - {filename}")
        
        row = 3
        
        # Estrutura de tópicos
        if 'topics_structure' in results:
            structure = results['topics_structure']
            
            ws_overview.cell(row=row, column=1, value="TÓPICOS PRINCIPAIS")
            self._style_subheader(ws_overview.cell(row=row, column=1))
            row += 1
            
            for topic in structure.get('main_topics', []):
                ws_overview.cell(row=row, column=1, value=f"• {topic}")
                row += 1
            
            row += 1
            
            # Hierarquia de tópicos
            if 'topic_hierarchy' in structure:
                ws_overview.cell(row=row, column=1, value="ESTRUTURA HIERÁRQUICA")
                self._style_subheader(ws_overview.cell(row=row, column=1))
                row += 1
                
                for main_topic, subtopics in structure['topic_hierarchy'].items():
                    ws_overview.cell(row=row, column=1, value=f"{main_topic}:")
                    self._style_cell(ws_overview.cell(row=row, column=1), bold=True)
                    row += 1
                    
                    for subtopic in subtopics:
                        ws_overview.cell(row=row, column=2, value=f"- {subtopic}")
                        row += 1
                    
                    row += 1
        
        ws_overview.column_dimensions['A'].width = 40
        ws_overview.column_dimensions['B'].width = 40
        
        # Sheet 2: Análise Detalhada por Tópico
        if 'topic_analyses' in results:
            ws_detailed = wb.create_sheet("Análise Detalhada")
            self._add_header(ws_detailed, "Análise Detalhada por Tópico")
            
            row = 3
            
            for topic_name, topic_data in results['topic_analyses'].items():
                # Nome do tópico
                ws_detailed.cell(row=row, column=1, value=topic_name.upper())
                self._style_subheader(ws_detailed.cell(row=row, column=1))
                row += 1
                
                # Resumo
                if 'summary' in topic_data:
                    ws_detailed.cell(row=row, column=1, value="Resumo:")
                    self._style_cell(ws_detailed.cell(row=row, column=1), bold=True)
                    ws_detailed.cell(row=row, column=2, value=topic_data['summary'])
                    ws_detailed.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
                    row += 1
                
                # Pontos-chave
                if 'key_points' in topic_data:
                    ws_detailed.cell(row=row, column=1, value="Pontos-chave:")
                    self._style_cell(ws_detailed.cell(row=row, column=1), bold=True)
                    row += 1
                    
                    for point in topic_data['key_points']:
                        ws_detailed.cell(row=row, column=2, value=f"• {point}")
                        row += 1
                
                row += 2  # Espaço entre tópicos
            
            ws_detailed.column_dimensions['A'].width = 20
            ws_detailed.column_dimensions['B'].width = 60
    
    def _create_data_extraction_sheets(self, wb: openpyxl.Workbook, results: Dict[str, Any], filename: str):
        """Cria sheets para extração de dados"""
        
        ws_data = wb.create_sheet("Dados Extraídos")
        self._add_header(ws_data, f"Dados Extraídos - {filename}")
        
        if 'extracted_data' in results:
            extracted = results['extracted_data']
            
            row = 3
            col = 1
            
            # Criar tabela para cada tipo de dado
            data_types = {
                'dates': 'DATAS',
                'numbers': 'NÚMEROS',
                'percentages': 'PERCENTUAIS',
                'currencies': 'VALORES MONETÁRIOS',
                'names': 'NOMES',
                'locations': 'LOCALIZAÇÕES',
                'organizations': 'ORGANIZAÇÕES'
            }
            
            for data_type, title in data_types.items():
                if data_type in extracted and extracted[data_type]:
                    # Cabeçalho da categoria
                    ws_data.cell(row=row, column=col, value=title)
                    self._style_subheader(ws_data.cell(row=row, column=col))
                    row += 1
                    
                    # Dados
                    for item in extracted[data_type]:
                        ws_data.cell(row=row, column=col, value=item)
                        row += 1
                    
                    row += 2  # Espaço entre categorias
                    
                    # Próxima coluna se muitos itens
                    if row > 30:
                        col += 2
                        row = 3
            
            # Ajustar larguras
            for i in range(1, col + 1):
                ws_data.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 30
    
    def _create_clause_analysis_sheets(self, wb: openpyxl.Workbook, results: Dict[str, Any], filename: str):
        """Cria sheets para análise de cláusulas"""
        
        ws_clauses = wb.create_sheet("Análise de Cláusulas")
        self._add_header(ws_clauses, f"Análise de Cláusulas - {filename}")
        
        if 'clauses' in results:
            row = 3
            
            for i, clause_data in enumerate(results['clauses'], 1):
                # Seção
                ws_clauses.cell(row=row, column=1, value=f"SEÇÃO {clause_data.get('section', i)}")
                self._style_subheader(ws_clauses.cell(row=row, column=1))
                row += 1
                
                # Cláusulas encontradas
                if 'clauses_found' in clause_data:
                    ws_clauses.cell(row=row, column=1, value="Cláusulas:")
                    self._style_cell(ws_clauses.cell(row=row, column=1), bold=True)
                    row += 1
                    
                    for clause in clause_data['clauses_found']:
                        ws_clauses.cell(row=row, column=2, value=f"• {clause}")
                        row += 1
                
                # Termos-chave
                if 'key_terms' in clause_data:
                    ws_clauses.cell(row=row, column=1, value="Termos-chave:")
                    self._style_cell(ws_clauses.cell(row=row, column=1), bold=True)
                    row += 1
                    
                    for term in clause_data['key_terms']:
                        ws_clauses.cell(row=row, column=2, value=f"• {term}")
                        row += 1
                
                # Obrigações
                if 'obligations' in clause_data:
                    ws_clauses.cell(row=row, column=1, value="Obrigações:")
                    self._style_cell(ws_clauses.cell(row=row, column=1), bold=True)
                    row += 1
                    
                    for obligation in clause_data['obligations']:
                        ws_clauses.cell(row=row, column=2, value=f"• {obligation}")
                        row += 1
                
                row += 2  # Espaço entre seções
        
        ws_clauses.column_dimensions['A'].width = 20
        ws_clauses.column_dimensions['B'].width = 60
    
    def _create_timeline_sheets(self, wb: openpyxl.Workbook, results: Dict[str, Any], filename: str):
        """Cria sheets para timeline de eventos"""
        
        ws_timeline = wb.create_sheet("Timeline")
        self._add_header(ws_timeline, f"Timeline de Eventos - {filename}")
        
        if 'timeline' in results:
            # Cabeçalhos da tabela
            headers = ['Data', 'Evento', 'Importância']
            row = 3
            
            for col, header in enumerate(headers, 1):
                ws_timeline.cell(row=row, column=col, value=header)
                self._style_subheader(ws_timeline.cell(row=row, column=col))
            
            row += 1
            
            # Eventos
            for event in results['timeline']:
                ws_timeline.cell(row=row, column=1, value=event.get('date', ''))
                ws_timeline.cell(row=row, column=2, value=event.get('event', ''))
                ws_timeline.cell(row=row, column=3, value=event.get('importance', ''))
                
                # Colorir baseado na importância
                importance = event.get('importance', '').lower()
                if importance == 'alta':
                    self._style_cell(ws_timeline.cell(row=row, column=3), bg_color='FFF2C5C5')
                elif importance == 'média':
                    self._style_cell(ws_timeline.cell(row=row, column=3), bg_color='FFFFF2CC')
                
                row += 1
        
        # Ajustar larguras
        ws_timeline.column_dimensions['A'].width = 15
        ws_timeline.column_dimensions['B'].width = 50
        ws_timeline.column_dimensions['C'].width = 15
    
    def _create_general_analysis_sheets(self, wb: openpyxl.Workbook, results: Dict[str, Any], filename: str, analysis_type: str):
        """Cria sheets para análise geral"""
        
        ws_general = wb.create_sheet("Análise")
        self._add_header(ws_general, f"{analysis_type} - {filename}")
        
        row = 3
        
        # Resultado da análise
        if 'results' in results:
            ws_general.cell(row=row, column=1, value="RESULTADO DA ANÁLISE")
            self._style_subheader(ws_general.cell(row=row, column=1))
            row += 1
            
            # Se o resultado for texto simples
            if isinstance(results['results'], str):
                ws_general.cell(row=row, column=1, value=results['results'])
                ws_general.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            
        ws_general.column_dimensions['A'].width = 80
    
    def _add_header(self, ws: openpyxl.worksheet.worksheet.Worksheet, title: str):
        """Adiciona cabeçalho padrão à planilha"""
        
        # Título
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(size=16, bold=True, color='FFFFFF')
        ws.cell(row=1, column=1).fill = PatternFill(start_color=self.colors['header'], 
                                                   end_color=self.colors['header'], 
                                                   fill_type='solid')
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        
        # Data de geração
        ws.cell(row=2, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        ws.cell(row=2, column=1).font = Font(size=10, italic=True)
        
        # Mesclar células do título
        ws.merge_cells('A1:D1')
    
    def _style_subheader(self, cell):
        """Aplica estilo de subcabeçalho"""
        cell.font = Font(size=12, bold=True, color='FF000000')
        cell.fill = PatternFill(start_color=self.colors['subheader'], 
                               end_color=self.colors['subheader'], 
                               fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def _style_cell(self, cell, bold=False, bg_color=None):
        """Aplica estilo personalizado à célula"""
        if bold:
            cell.font = Font(bold=True)
        
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
